from django.test import TestCase

# Create your tests here.
# a = 'abc'
# a = a.encode('utf-8')
# print(a)
import os
# import hashlib
# h = hashlib.md5()
#
# save_hash = ''
# save_path = r'D:\迅雷下载\cn_windows_7_ultimate_x86_dvd_x15-65907.iso'
# with open(save_path, 'rb') as f:
#     start = hashlib.md5()
#     for i in range(10):
#         file = f.read(50 * 1024 * 1024)
#         start.update(file)
#         save_hash = start.hexdigest()
#     print(save_hash)
#     end = hashlib.md5()
#     size = len(f.read())
#     seek_size = size - 500 * 1024 * 1024
#     if seek_size > 0:
#         f.seek(seek_size)
#         for i in range(10):
#             file = f.read(50 * 1024 * 1024)
#             end.update(file)
#             save_hash += end.hexdigest()
#     else:
#         save_hash += save_hash

# print(save_hash)

# with open(r'D:\迅雷下载\cn_windows_7_ultimate_x86_dvd_x15-65907.iso', 'rb') as f:
#     for i in range(10):
#         file = f.read(50*1024*1024)
#         h.update(file)
# print(h.hexdigest())
#
# ha = hashlib.md5()
# with open(r'D:\迅雷下载\cn_windows_7_ultimate_x86_dvd_x15-65907.iso', 'rb') as f:
#     size = os.path.getsize(r'D:\迅雷下载\cn_windows_7_ultimate_x86_dvd_x15-65907.iso')
#     seek_size = size - 500*1024*1024
#     f.seek(seek_size)
#     for i in range(10):
#         file = f.read(50*1024*1024)
#         ha.update(file)
# print(ha.hexdigest())


# while len(file) > 0:
#     # print(len(file))
#     h.update(file)
#     print(h.hexdigest())
# f.seek(5 * 1024 * 1024)
# file = f.read(5*1024*1024)
# h.update(file)
# print(h.hexdigest())

# f.seek(size - 5 * 1024 * 1024)
# file = f.read(5 * 1024 * 1024)
# h.update(file)
# print(h.hexdigest())
# f = open(r'D:\迅雷下载\PySide-1.2.2.win32-py3.4.exe', 'rb')
# 4d32e93aa75aec7753da0317b8db6dca
# file = f.read(5*1024*1024)


# print(h.hexdigest())
# h.update(file)
# import time
# path = r'C:\Users\yx\Desktop\慈铭月报表'
# for d in os.listdir(path):
#     file_path = os.path.join(path, d)
#     ctime = os.path.getctime(file_path)
#     if time.time() - ctime > 60 * 60 * 24:
#         print(ctime, d)
# import re
# a = '\/:*?"<>'
# s = 'abcs?d\\'
# print(re.search(r'[\\/:*?"<>]', s))

# os.rename(r'C:\Users\yx\Desktop\闪电音频格式转换器11', r'C:\Users\yx\Desktop\闪电音频格式转换器')

# with open('../visit.log', 'a+') as f:
#     f.write('abc')
# path = r'F:\game\COD11'
# import hashlib
#
# for file in os.listdir(path):
#     file_path = os.path.join(path, file)
#     if os.path.isfile(file_path):
#         with open(file_path, 'rb') as f:
#             fi = f.read()
#             h = hashlib.md5()
#             h.update(fi)
#             print(file, h.hexdigest())

# with open(r'F:\game\COD11\COD11.part02.rar', 'rb') as f:
#     fi = f.read()
#     h.update(fi)
#     print(h.hexdigest())

import shutil

# a, b = json.loads('["",""]')
# print(os.listdir(r'D:\迅雷下载'))
# print(os.path.join(['2017级大理大学毕业环节资料填写指南', '彩超', '神圣']))

# shutil.move(r'C:\Users\yx\Desktop\项目图片\新建文件夹', r'C:\Users\yx\Desktop\项目图片')
# # print(os.path.split(r'C:\Users\yx\Desktop\项目图片\新建文件夹'))
#
# print(r'C:\Users\yx\Desktop\项目图片' in r'C:\Users\yx\Desktop\项目图片\新建文件夹')
# a = r'D:\python\django\NAS\staticfile\NAS\1\测试\临床医学 专科'
# b = r'D:\python\django\NAS\staticfile\NAS\1'
# print(a.strip(b))
# shutil.copyfile()
# shutil.copy(r'D:\python\django\NAS\staticfile\NAS\1\测试\临床医学 本科', r'D:\python\django\NAS\staticfile\NAS\1\2017级大理大学毕业环节资料填写指南')

'2818072576'
# path = r'D:\python\django\NAS\staticfile\NAS\1'
# # print(os.path.getsize()
# print(os.walk(path))
a = 1551051511
# print(a + ~a)
# path = r'G:\NAS\staticfile\NAS\杨旭'
path = r'E:\婚纱照'


# b = 0
# def fun(path):
#     global b
#     for d in os.listdir(path):
#         dd = os.path.join(path, d)
#         if os.path.isdir(dd):
#             fun(dd)
#         else:
#             b += os.path.getsize(dd)
#             # print(os.path.getsize(dd))
#
# a = 'width="553px"'
# import re
# print(re.sub('width="\d+px"', a, 'width="100%"'))
#
# import pypptx

# fun(path)
# print(b)
# a = r'G:\NAS\staticfile\NAS\杨旭'
# print(r'G:\NAS\staticfile\NAS\杨旭\照片\百度网盘qq下载\MjAxNF8wN18xMV8xNl80M180My5qcGc='.strip(a))


def ppt_to_pdf(filepath, savepath):
    # ppt和pptx文件转换
    from win32com.client import gencache
    if os.path.isfile(savepath):
        return 2, '已经转换了'
    p = gencache.EnsureDispatch("PowerPoint.Application")
    try:
        ppt = p.Presentations.Open(filepath, False, False, False)
    except Exception as e:
        return os.path.split(filepath)[1], "转化失败,失败原因%s" % e
    ppt.ExportAsFixedFormat(savepath, 2, PrintRange=None)
    p.Quit()
    return 1, '转换成功'
# D:\python\django\NAS\staticfile\NAS\1\576O5bm05YGl5bq357O757uf5Y,K6K6,5aSH6L,e5o6lMjAxODAxMDIucHB0eA== D:\python\django\NAS\staticfile\buffer\DOC\美年健康系统及设备连接20180102.pdf
# ppt_to_pdf(r'D:\python\django\NAS\staticfile\NAS\1\576O5bm05YGl5bq357O757uf5Y,K6K6,5aSH6L,e5o6lMjAxODAxMDIucHB0eA==', r'D:\python\django\NAS\staticfile\buffer\DOC\美年健康系统及设备连接20180102.pdf')


def txt_to_html(filepath, savepath):
    # with open(filepath, 'r', encoding='gbk') as f:
    #     file = f.read()
    # with open(r'D:\python\django\NAS\111.txt', 'w', encoding='utf-8') as f:
    #     f.write(file)

    # with open(filepath, "r", encoding='gbk') as file:
    #     contents = file.readlines()
    # #
    # with open(savepath, "w", encoding='gbk') as e:
    #     for lines in contents:
    #         e.write("<pre>" + lines + "</pre>\n")

    with open(filepath, 'rb') as f:
        data = f.read()
        import chardet
        print(chardet.detect(data)['encoding'])

# txt_to_html(r'D:\python\django\NAS\Game.ini', r'D:\python\django\NAS\templates\txt.html')
def excel_to_html():
    import pandas as pd
    import codecs
    pd.set_option('display.width', 1000)
    pd.set_option('colheader_justify', 'center')
    xd = pd.ExcelFile(r'C:\Users\yx\Desktop\项目图片\新建文件夹\111.xlsx')
    df = xd.parse()
    for key in df.keys():
        df[key] = df[key].apply(lambda x: '' + str(x))
    pd.set_option('colheader_justify', 'center')
    #设置html文件格式
    html_string = '''
    <!DOCTYPE html>  
        <head>  
        <meta charset="UTF-8">  
        <title></title>  
        </head>  
        <link rel="stylesheet" type="text/css" href="df_style.css"/>  
        <body>
            {table}  
        </body>
    </html>.'''
    with open(r'C:\Users\yx\Desktop\项目图片\新建文件夹\云南民族大学附属中学.html', encoding='utf-8', mode='w') as f:
        f.write(html_string.format(table=df.to_html()))



def test01():
    import pandas
    pd = pandas.ExcelFile(r'C:\Users\yx\Desktop\项目图片\新建文件夹\云南民族大学附属中学.xls')
    pd = pd.parse()
    print(pd)
    for key in pd.keys():
        pd[key] = pd[key].apply(lambda x: str(x))
    # print(pd.to_html())

# excel_to_html()
# test01()

def to_html(excel_list):
    html = ''
    for sheet in excel_list:
        sheet_name = list(sheet.keys())[0]
        sheet_text = '<h1>{0}</h1>\n<table class="table table-bordered table-hover {0}">\n\t<tbody>'.format(sheet_name)
        value_list = sheet[sheet_name]
        for row, col_list in enumerate(value_list):
            tr = '<tr>\n'
            for col, value in enumerate(col_list):
                td = '\t<td>{}</td>\n'.format(value)
                tr += td
            tr += '</tr>\n'
            sheet_text += tr
        sheet_text += '\t</tbody>\n</table>\n'
        html += sheet_text
    return html


def read_excel(filepath, filename):
    save_path = os.path.join(settings.BUFFER_DIR, 'DOC', filename)
    shutil.copy(filepath, save_path)
    filepath = save_path
    if filename[-4:] == '.xls':
        print('调用xlrd读取excel文件', filepath)
        excel = xlrd.open_workbook(filepath)
        print(excel)
    elif filename[-5:] == '.xlsx':
        print('调用openpyxl读取excel文件')
        excel = openpyxl.open(filepath, data_only=True)
        excel_list = []
        merged_cell_dict = {}
        for sheet in excel.sheetnames:
            wb = excel[sheet]
            merged_list = []
            max_rows = wb.max_row
            max_cols = wb.max_column
            for merged in wb.merged_cell_ranges:
                left, right = str(merged).split(':')
                re_left = re.match('[A-Z]+', left).group()
                re_right = re.match('[A-Z]+', right).group()
                merged = '{}:{}'.format(
                    left.replace(re_left, str(column_index_from_string(re_left)) + '-'),
                    right.replace(re_right, str(column_index_from_string(re_right)) + '-'))
                merged_list.append(merged)
            merged_cell_dict[sheet] = merged_list
            sheet_dict = {}
            sheet_list = []
            for row in range(1, max_rows+1):
                row_list = []
                for col in range(1, max_cols+1):
                    val = wb.cell(row, col).value
                    if val:
                        row_list.append(val)
                    else:
                        row_list.append('')
                sheet_list.append(row_list)
                sheet_dict[sheet] = sheet_list
            excel_list.append(sheet_dict)
        return excel_list, merged_cell_dict


# excel_list, merged_cell_dict = read_excel(r'C:\Users\yx\Desktop\项目图片\新建文件夹\美年大健康账号信息.xlsx')
# html = to_html(excel_list)
# print(merged_cell_dict)
def sort_merged_dict(merged):
    row_merged_dict = {}
    col_merged_dict = {}
    for key in merged.keys():
        value = merged[key]
        row_merged_dict[key] = []
        col_merged_dict[key] = []
        for val in value:
            left_col = val.split(':')[0].split('-')[0]
            left_row = val.split(':')[0].split('-')[1]
            rig_col = val.split(':')[1].split('-')[0]
            rig_row = val.split(':')[1].split('-')[1]
            if left_row == rig_row:
                row_merged_dict[key].append({left_row: [left_col, rig_col]})
                # print('行合并：将第{}行从第{}列到第{}列进行合并'.format(left_row, left_col, rig_col))
            if left_col == rig_col:
                col_merged_dict[key].append({left_col: [left_row, rig_row]})
                # print('列合并：将第{}列从第{}行到第{}行进行合并'.format(left_col, left_row, rig_row))
    print(row_merged_dict)
    print(col_merged_dict)
    return row_merged_dict, col_merged_dict

# ttt(merged_cell_dict)
# print('<td {}>{}</td>'.format('1', '2'))
def read_xlrd(filepath):
    import xlrd, openpyxl, re
    from openpyxl.utils import column_index_from_string
    excel = xlrd.open_workbook(filepath, formatting_info=True)
    # excel_list = []
    # merged_cell_dict = {}
    # for sheet in wb.sheet_names():
    #     ws = wb.sheet_by_name(sheet)
    #     print(ws.nrows, ws.ncols)
    #     print()
    # print('调用openpyxl读取excel文件')
    # excel = openpyxl.open(filepath, data_only=True)
    excel_list = []
    merged_cell_dict = {}
    for sheet in excel.sheet_names():
        wb = excel[sheet]
        merged_list = []
        max_rows = wb.nrows
        max_cols = wb.ncols
        print(wb.merged_cells)
        # for merged in wb.merged_cell_ranges:
        #     left, right = str(merged).split(':')
        #     re_left = re.match('[A-Z]+', left).group()
        #     re_right = re.match('[A-Z]+', right).group()
        #     merged = '{}:{}'.format(
        #         left.replace(re_left, str(column_index_from_string(re_left)) + '-'),
        #         right.replace(re_right, str(column_index_from_string(re_right)) + '-'))
        #     merged_list.append(merged)
        merged_cell_dict[sheet] = wb.merged_cells
        sheet_dict = {}
        sheet_list = []
        for row in range(max_rows):
            row_list = []
            for col in range(max_cols):
                val = wb.cell(row, col).value
                if val:
                    row_list.append(val)
                else:
                    row_list.append('')
            sheet_list.append(row_list)
            sheet_dict[sheet] = sheet_list
        excel_list.append(sheet_dict)
    return excel_list#, merged_cell_dict

# {'Sheet1': ['1-2:3-2', '5-4:5-7']}
#            [(1, 2, 0, 3), (3, 7, 4, 5)]
# print(read_xlrd(r'C:\Users\yx\Desktop\项目图片\新建文件夹\新建 XLSX 工作表.xls'))
# import win32com.client as win32
#
# fname = r'C:\Users\yx\Desktop\项目图片\新建文件夹\新建 XLSX 工作表.xls'
# excel = win32.gencache.EnsureDispatch('Excel.Application')
# wb = excel.Workbooks.Open(fname)
#
# wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
# wb.Close()                               #FileFormat = 56 is for .xls extension
# excel.Application.Quit()
for i in range(1, 10):
    print(i)


