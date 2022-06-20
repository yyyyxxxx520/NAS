from django.shortcuts import render, redirect
from django.http import JsonResponse, FileResponse, HttpResponse, StreamingHttpResponse
from django.conf import settings
from django.utils.encoding import escape_uri_path
from wsgiref.util import FileWrapper
import os, time, base64, shutil, re
from nas_app.DataEncryption import Encrypt
import hashlib
import json
import sqlite3
import zipfile
import subprocess
import platform
import ctypes
import pythoncom
import chardet
from win32com.client import gencache
import win32com.client as win32
import openpyxl
from openpyxl.utils import column_index_from_string



# Create your views here.


class NASSqlite3:
    def __init__(self, user=0):
        self.NAS_DB = sqlite3.connect(os.path.join(settings.BASE_DIR, 'NAS.sqlite3'))
        self.NAS_CU = self.NAS_DB.cursor()
        self.nas_user = user

    # 验证登录用户信息
    def login(self, username, password):
        sql = """select * from user_info where name='{}' and password='{}';""".format(username, password)
        self.NAS_CU.execute(sql)
        result = self.NAS_CU.fetchone()
        return result

    # 验证用户注册信息
    def register(self, username, sex=None, password=None, isExist=False):
        try:
            if isExist:
                sql = """select * from user_info where name='{}';""".format(username)
                self.NAS_CU.execute(sql)
                return self.NAS_CU.fetchone()
            else:
                sql = """insert into user_info (name, sex, password)
                    values('{}', '{}', '{}');""".format(username, sex, password)
                self.NAS_CU.execute(sql)
                self.NAS_DB.commit()
                return 1
        except Exception as e:
            return e

    # 查询数据库中用户已用空间和总的空间
    def NAS_select_user_capacity(self):
        username = Encrypt().encipher(self.nas_user, settings.USERNAME_KEY)
        sql = """select capacityUsed,capacityTotal from user_info where name='{}';""".format(username)
        self.NAS_CU.execute(sql)
        return self.NAS_CU.fetchone()

    # 修改数据库中用户已用空间
    def NAS_update_user_capacity(self, change_size):
        try:
            username = Encrypt().encipher(self.nas_user, settings.USERNAME_KEY)
            select_sql = """select capacityUsed from user_info where name='{}';""".format(username)
            self.NAS_CU.execute(select_sql)
            capacityUsed = self.NAS_CU.fetchone()[0]
            capacityUsed += change_size
            update_sql = """update user_info set capacityUsed='{}' where name='{}';""".format(capacityUsed, username)
            self.NAS_CU.execute(update_sql)
            self.NAS_DB.commit()
            return [1, '修改成功']
        except Exception as e:
            return [0, str(e)]

    # 将上传的文件信息插入数据库中
    def NAS_file_insert_sql(self, file_name, file_base64, file_path, file_size, file_type, file_class, upload_date,
                            file_state='normal'):
        try:
            file_path = file_path.replace('&', '-')
            file_size = file_size_format(file_size)
            # print(file_name, file_base64, file_path, file_size, file_type, file_class, upload_date)
            sql = """insert into file_info 
                        (file_name, file_base64, file_path, file_size, file_type,file_class, upload_date, file_state, nas_user)
                                values('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}');""".format(
                file_name, file_base64, file_path, file_size, file_type, file_class, upload_date, file_state,
                self.nas_user)
            self.NAS_CU.execute(sql)
            self.NAS_DB.commit()
            return [1, '插入成功']
        except Exception as e:
            return [0, '错误信息：{}'.format(e)]

    # 输入文件名、base64名、文件路径修改文件的其他信息
    def NAS_file_update_sql(self, file_name, file_base64, file_path, file_size, file_type, file_class, upload_date,
                            file_state='normal'):
        try:
            file_path = file_path.replace('&', '-')
            file_size = file_size_format(file_size)
            sql = """update file_info set file_size = '{}', file_type='{}', file_class='{}',upload_date='{}',file_state='{}'
             where file_name = '{}' and file_base64='{}' and file_path='{}' and nas_user='{}';""".format(file_size,
                                                                                                         file_type,
                                                                                                         file_class,
                                                                                                         upload_date,
                                                                                                         file_state,
                                                                                                         file_name,
                                                                                                         file_base64,
                                                                                                         file_path,
                                                                                                         self.nas_user)
            self.NAS_CU.execute(sql)
            self.NAS_DB.commit()
            return [1, '更新成功']
        except Exception as e:
            return [0, '错误信息：{}'.format(e)]

    # 修改数据库中的文件名
    def NAS_file_rename_sql(self, old_file_name, old_file_base64, new_file_name, new_file_base64, file_path):
        try:
            file_path = file_path.replace('&', '-')
            sql = '''update file_info set file_name="{}", file_base64="{}"
                    where file_name="{}" and file_base64="{}" and file_path="{}" and nas_user="{}";
            '''.format(new_file_name, new_file_base64, old_file_name, old_file_base64, file_path, self.nas_user)
            self.NAS_CU.execute(sql)
            self.NAS_DB.commit()
            return [1, '修改成功']
        except Exception as e:
            return [0, '错误信息：{}'.format(e)]

    # 更改文件夹名称后修改数据库中的文件路径
    def NAS_dir_rename_sql(self, old_file_path, new_file_path):
        try:
            sql = """select DISTINCT file_path from file_info where file_path like '{}%' and nas_user="{}";""".format(
                old_file_path, self.nas_user)
            self.NAS_CU.execute(sql)
            result = self.NAS_CU.fetchall()
            for cur in result:
                old_path = cur[0]
                new_path = old_path.replace(old_file_path, new_file_path)
                sql = 'update file_info set file_path = "{}" where file_path = "{}" and nas_user="{}";'.format(new_path,
                                                                                                               old_path,
                                                                                                               self.nas_user)
                self.NAS_CU.execute(sql)
            self.NAS_DB.commit()
            return [1, '修改成功']
        except Exception as e:
            return [0, e]

    # 移动文件后修改数据库中的文件路径
    def NAS_file_repath_sql(self, file_name, cur_dir, new_file_path):
        try:
            sql = 'update file_info set file_path = "{}" where file_path = "{}" and file_name="{}" and nas_user="{}";'.format(
                new_file_path, cur_dir, file_name, self.nas_user)
            self.NAS_CU.execute(sql)
            self.NAS_DB.commit()
            return [1, '修改成功']
        except Exception as e:
            return [0, e]

    # 复制文件夹和文件时将路径插入数据库中
    def NAS_copy_insert_sql(self, old_file_path, new_file_path, isDir, cur_dir=''):
        if isDir:
            select_dir_sql = """select * from file_info where file_path like '{}%' and nas_user='{}';""".format(
                old_file_path, self.nas_user)
            self.NAS_CU.execute(select_dir_sql)
            result = self.NAS_CU.fetchall()
            if result:
                for fid, name, base, path, size, ftype, fclass, date, state, user in result:
                    new_path = path.replace(old_file_path, new_file_path)
                    if self.NAS_select_filename(base, new_path) == 'None':
                        code, msg = self.NAS_file_insert_sql(name, base, new_path, size, ftype, fclass, get_time())
                        if not code:
                            return [code, msg]
                return [1, '数据库插入成功']
            return [0, '复制的文件不存在，请刷新后重试！']
        else:
            file_name = os.path.split(new_file_path)[1]
            new_file_path = new_file_path.strip(file_name).strip('\\')
            select_file_info_sql = """select * from file_info where file_base64='{}' and file_path='{}' and 
            nas_user='{}';""".format(file_name, cur_dir, self.nas_user)
            self.NAS_CU.execute(select_file_info_sql)
            result = self.NAS_CU.fetchone()
            if result:
                fid, name, base, path, size, ftype, fclass, date, state, user = result
                new_path = path.replace(old_file_path, new_file_path)
                if self.NAS_select_filename(base, new_path) == 'None':
                    code, msg = self.NAS_file_insert_sql(name, base, new_path, size, ftype, fclass, get_time())
                    if not code:
                        return [code, msg]
                return [1, '数据库插入成功']
            return [0, '复制的文件不存在，请刷新后重试！']

    # 输入参数，查询符合条件的数据
    def NAS_file_select_sql(self, file_name=None, file_path=None, file_type=None, file_class=None, upload_date=None,
                            file_state='normal', start=0):
        try:
            sql = """
                select file_name, file_base64, file_type,upload_date, file_size, file_path from file_info where file_state = '{}'
                and nas_user='{}'
                """.format(file_state, self.nas_user)
            if file_name:
                sql += """ and file_name = '{}'""".format(file_name)
            if not file_path is None:
                sql += """ and file_path = '{}'""".format(file_path)
            if file_type:
                file_type = str(file_type).replace("['", '').replace("']", '')
                sql += """ and file_type in ('{}')""".format(file_type)
            if file_class:
                sql += """ and file_class = '{}'""".format(file_class)
            if upload_date:
                sql += """ and upload_date = '{}'""".format(upload_date)
            sql += ' limit {},{};'.format(start, 50)
            self.NAS_CU.execute(sql)
            self.NAS_DB.commit()
            result = self.NAS_CU.fetchall()
            return [1, result]
        except Exception as e:
            return [0, '错误信息：{}'.format(e)]

    # 输入文件名称和路径查询文件的base64名称
    def NAS_select_savename(self, filename, cur_dir):

        sql = 'select file_base64 from file_info where file_name="{}" and file_path="{}" and nas_user="{}"; '.format(
            filename, cur_dir, self.nas_user)
        self.NAS_CU.execute(sql)
        result = self.NAS_CU.fetchone()
        if result:
            return result[0]
        return None

    # 输入文件的base64名称和路径查询文件名称
    def NAS_select_filename(self, file_base64, cur_dir):
        sql = 'select file_name from file_info where file_base64="{}" and file_path="{}" and nas_user="{}"; '.format(
            file_base64, cur_dir, self.nas_user)
        self.NAS_CU.execute(sql)
        result = self.NAS_CU.fetchone()
        if result:
            return result[0]
        else:
            return 'None'

    # 删除数据库中文件数据
    def NAS_delete_file(self, filename, upload_date, file_size, cur_dir):

        try:
            if filename and upload_date:
                save_name = self.NAS_select_savename(filename, cur_dir)
                sql = """delete from file_info
                where file_name='{}' and upload_date='{}' and file_size='{}' and file_path='{}' and nas_user='{}';""" \
                    .format(filename, upload_date, file_size, cur_dir, self.nas_user)
                self.NAS_CU.execute(sql)
                self.NAS_DB.commit()
                return [1, save_name]
            else:
                sql = """delete from file_info
                                where file_path = '{}' and nas_user='{}'""".format(cur_dir, self.nas_user)

                self.NAS_CU.execute(sql)
                self.NAS_DB.commit()
                return [1, cur_dir]
        except Exception as e:
            # print(e)
            return [0, e]
        # return [1, 'ces']

    # 查找数据库中符合条件的数据
    def NAS_file_search_sql(self, search_val, file_class):
        try:
            sql = """select file_name, file_base64, file_type,upload_date, file_size, file_path from file_info 
                    where file_name like '%{}%' and nas_user='{}'""".format(search_val, self.nas_user)
            if file_class:
                sql += " and file_type in ({});".format(file_class).replace('[', '').replace(']', '')
            # sql += 'limit {},{};'.format(start, start + 50)
            self.NAS_CU.execute(sql)
            result = self.NAS_CU.fetchall()
            return [1, result]

        except Exception as e:
            return [0, e]


def login(request):
    if request.META.get('HTTP_X_FORWARDED_FOR'):
        ip = request.META.get("HTTP_X_FORWARDED_FOR")
    else:
        ip = request.META.get("REMOTE_ADDR")
    UserAgent = request.headers.get('User-Agent')
    if request.method == 'GET':
        s = 'login：名称：{}，访问IP：{}，访问时间：{}\n'.format(UserAgent, ip, time.strftime("%Y-%m-%d %H:%M"))
        with open('visit.log', 'a') as f:
            f.write(s)
        return render(request, 'login.html')
    else:
        username = request.POST.get('username')
        password = request.POST.get('password')
        if username and password:
            username = Encrypt().encipher(username, settings.USERNAME_KEY)
            password = Encrypt().encipher(password, settings.USERNAME_KEY)
            result = NASSqlite3().login(username, password)
            if result:
                name = request.POST.get('username')
                request.session['uid'] = result[0]
                request.session['user'] = name
                request.session['sex'] = result[2]
                request.session['nas_dir'] = os.path.join(settings.NAS_DIR, name)
                if not os.path.exists(os.path.join(settings.NAS_DIR, name)):
                    os.mkdir(os.path.join(settings.NAS_DIR, name))
                if not os.path.exists(settings.BUFFER_DIR):
                    os.mkdir(settings.BUFFER_DIR)

                s = '登录成功：用户名：{}，名称：{}，访问IP：{}，访问时间：{}\n'.format(name, UserAgent, ip, time.strftime("%Y-%m-%d %H:%M"))
                with open('login.log', 'a') as f:
                    f.write(s)
                return JsonResponse({'code': 1, 'url': '/'})
            else:
                return JsonResponse({'code': 0, 'url': '/login/', 'mes': '用户名密码错误请重新输入'}, charset='utf-8')
        else:
            return JsonResponse({'code': 0, 'url': '/login/', 'mes': '请输入用户名和密码'}, charset='utf-8')


def register(request):
    if request.method == 'GET':
        if request.META.get('HTTP_X_FORWARDED_FOR'):
            ip = request.META.get("HTTP_X_FORWARDED_FOR")
        else:
            ip = request.META.get("REMOTE_ADDR")
        UserAgent = request.headers.get('User-Agent')
        s = 'register：名称：{}，访问IP：{}，访问时间：{}\n'.format(UserAgent, ip, time.strftime("%Y-%m-%d %H:%M"))
        with open('visit.log', 'a') as f:
            f.write(s)
        return render(request, 'register.html')
    else:
        username = request.POST.get('username')
        sex = request.POST.get('sex')
        password = request.POST.get('password')
        confirm = request.POST.get('confirm')
        authorization = request.POST.get('authorization')
        if username and sex and password and confirm:
            if password == confirm:
                if sex == '男' or sex == '女':
                    encipher_username = Encrypt().encipher(username, settings.USERNAME_KEY)
                    encipher_password = Encrypt().encipher(password, settings.USERNAME_KEY)
                    isExist = NASSqlite3().register(username=encipher_username, isExist=True)
                    if authorization != 'yangxu520':
                        return JsonResponse({'code': '0', 'url': '/register/', 'mes': '授权码不正确！！！'}, charset='utf-8')
                    if isExist:
                        return JsonResponse({'code': '0', 'url': '/register/', 'mes': '该用户名已被占用，请更改'}, charset='utf-8')
                    else:
                        result = NASSqlite3().register(encipher_username, sex, encipher_password)
                        if result == 1:
                            os.mkdir(os.path.join(settings.NAS_DIR, username))
                            return JsonResponse({'code': '1', 'url': '/login/', 'mes': '注册成功，将跳转至登陆界面'})
                        else:
                            return JsonResponse({'code': '0', 'url': '/register/', 'mes': str(result)})
        return JsonResponse({'code': '0', 'url': '/register/', 'mes': '注册错误，请核对'}, charset='utf-8')


def NAS(request):
    nas_dir = request.session.get('nas_dir')
    if request.method == 'GET':
        cur_dir = request.GET.get('cur_dir', '').strip('\\')
        state = request.GET.get('state', None)
        standby = request.GET.get('standby', None)
        sort = request.GET.get('sort', '')
        start = request.GET.get('start', 0)
        # 如果没有该文件夹，则直接跳转至首页
        if not os.path.exists(os.path.join(nas_dir, cur_dir)):
            if request.path == '/' or request.path == '/NAS/':
                return redirect('/login/')
            return redirect('/NAS/')
        # 查看符合类型的文件
        if state in settings.NAS_FILE_CLASS.keys():
            code, data = NASSqlite3(request.session.get('user')).NAS_file_select_sql(
                file_type=settings.NAS_FILE_CLASS[state], start=start)
            return JsonResponse({'code': code, 'folders': [], 'file_data': data, 'cur_dir': ''}, charset='utf-8')
        # 查看共享文件方法
        elif state == 'share':
            code, file_data = NASSqlite3(request.session.get('user')).NAS_file_select_sql(file_type='share',
                                                                                          start=start)
            return JsonResponse({'code': code, 'folders': [], 'file_data': file_data, 'cur_dir': ''}, charset='utf-8')
        # 查看隐私文件方法
        elif state == 'privacy':
            code, file_data = NASSqlite3(request.session.get('user')).NAS_file_select_sql(file_type='privacy',
                                                                                          start=start)
            return JsonResponse({'code': code, 'folders': [], 'file_data': file_data, 'cur_dir': ''}, charset='utf-8')
        # 查找数据方法
        elif state == 'search':
            search_val, file_class = json.loads(request.GET.get('standby', '["",""]'))
            folders = []
            # 查找所有文件
            if file_class == 'home':
                file_class = None

                # 递归查询与输入参数类似的文件夹名称
                def fun(path):
                    for file in os.listdir(path):
                        file_path = os.path.join(path, file)
                        if os.path.isdir(file_path):
                            if search_val in file:
                                fol_info = {'name': file,
                                            'fol_path': file_path.replace(nas_dir + '\\', ''),
                                            'ctime': get_time(os.path.getctime(file_path), cn=False),
                                            'mtime': get_time(os.path.getmtime(file_path), cn=False)}
                                folders.append(fol_info)
                                # folders.append(file_path.replace(request.session.get('nas_dir') + '\\', ''))#
                            fun(file_path)

                fun(nas_dir)
            # 查找指定类型的文件
            else:
                file_class = settings.NAS_FILE_CLASS.get(file_class, ['-'])
            code, data = NASSqlite3(request.session.get('user')).NAS_file_search_sql(search_val, file_class)
            if code == 0:
                return JsonResponse({'code': 0, 'folders': [], 'file_data': [], 'cur_dir': cur_dir}, charset='utf-8')
            return JsonResponse({'code': 1, 'folders': folders, 'file_data': data, 'cur_dir': cur_dir}, charset='utf-8')
        else:
            if cur_dir.split('\\')[0]:
                dir_split = cur_dir.split('\\')
            else:
                dir_split = None
            code, file_data = NASSqlite3(request.session.get('user')).NAS_file_select_sql(file_path=cur_dir,
                                                                                          start=start)
            if code == 0:
                return JsonResponse({'code': 0, 'folders': [], 'file_data': [], 'cur_dir': cur_dir}, charset='utf-8')
            folders = get_folders(nas_dir, cur_dir)
            if standby:
                return JsonResponse({'code': 1, 'folders': folders, 'file_data': file_data, 'cur_dir': cur_dir}, charset='utf-8')

            result = NASSqlite3(request.session.get('user')).NAS_select_user_capacity()
            capacityRatio = '%.1f' % (result[0] / result[1] * 100) + '%'
            capacityUsed = file_size_format(result[0])
            capacityTotal = file_size_format(result[1])
            return render(request, 'NAS.html',
                          {'user_name': request.session.get('user'), 'folders': folders, 'file_data': file_data,
                           'cur_dir': cur_dir, 'dir_split': dir_split, 'capacityUsed': capacityUsed,
                           'capacityTotal': capacityTotal,
                           'capacityRatio': capacityRatio})
    else:
        state = request.POST.get('state', None)
        cur_dir = request.POST.get('cur_dir', '').strip('\\')
        if state == 'new_folder':
            folder_name = request.POST.get('folder_name', None)
            if os.path.isdir(os.path.join(nas_dir, cur_dir)):
                if folder_name and validity(folder_name):
                    folder_path = os.path.join(nas_dir, cur_dir, folder_name)
                    folder_path = folder_path.replace('&', '-')
                    if os.path.exists(folder_path):
                        return JsonResponse({'code': 0, 'msg': '文件夹已存在，请重新输入！'}, charset='utf-8')
                    elif len(folder_path) > 247:
                        return JsonResponse({'code': 0, 'msg': '文件夹所在路径过长，无法创建!'}, charset='utf-8')
                    try:
                        os.mkdir(folder_path)
                        return JsonResponse({'code': 1, 'msg': '创建成功'})
                    except Exception as e:
                        return JsonResponse({'code': 0, 'msg': '创建文件夹时发生错误，错误信息：{}'.format(e)}, charset='utf-8')
                else:
                    return JsonResponse({'code': 0, 'msg': '名称不合法！'}, charset='utf-8')
            return JsonResponse({'code': 0, 'msg': '路径不合法！'}, charset='utf-8')
        elif state == 'rename':
            rename_dir = request.POST.get('dir', None)
            rename_file = request.POST.get('file', None)
            rename = request.POST.get('rename', None)
            if rename_dir:
                dir_path = os.path.join(nas_dir, cur_dir)
                dir_name = dir_path.rsplit('\\').pop(-1)
                rename_path = dir_path.replace(dir_name, rename)
                if os.path.isdir(dir_path) and not os.path.isdir(rename_path):
                    if rename and validity(rename):
                        try:
                            code, msg = NASSqlite3(request.session.get('user')).NAS_dir_rename_sql(cur_dir,
                                                                                                   cur_dir.replace(dir_name, rename))
                            if code:
                                os.rename(dir_path, rename_path)
                                return JsonResponse({'code': 1, 'msg': '修改成功'})
                            return JsonResponse({'code': 0, 'msg': '修改文件夹名称时发生错误，错误信息：{}'.format(msg)})
                        except Exception as e:
                            return JsonResponse({'code': 0, 'msg': '修改文件夹名称时发生错误，错误信息：{}'.format(e)})
                    return JsonResponse({'code': 0, 'msg': '新文件夹名不符合要求，请查证后再试！'})
                return JsonResponse({'code': 0, 'msg': '待修改的文件夹不存在或新文件夹名称重复，请查证后再试！'})
            elif rename_file:
                print(rename_file, cur_dir)
                old_save_name = NASSqlite3(request.session.get('user')).NAS_select_savename(rename_file, cur_dir)

                old_path = os.path.join(nas_dir, cur_dir, old_save_name)
                if os.path.isfile(old_path):
                    # 对文件名和base64名称进行重复校验，获取校验后的名称
                    save_path, file_name, save_name = name_nodup(os.path.join(nas_dir, cur_dir), rename,
                                                                 False,
                                                                 settings.IS_BASE64_ENCRYPT)
                    code, msg = NASSqlite3(request.session.get('user')).NAS_file_rename_sql(rename_file, old_save_name,
                                                                                            file_name, save_name,
                                                                                            cur_dir)
                    if code:
                        try:
                            os.rename(old_path, save_path)
                            return JsonResponse({'code': 1, 'msg': '修改成功！'})
                        except Exception as e:
                            return JsonResponse({'code': 0, 'msg': '修改文件名称时发生错误，错误信息：{}'.format(e)})
                    return JsonResponse({'code': code, 'msg': msg})
                return JsonResponse({'code': 0, 'msg': '文件修改失败，请刷新后重试'})
        elif state == 'get_folder':
            folder = request.POST.get('folder', '')
            cur_dir = os.path.join(cur_dir, folder)
            path = os.path.join(nas_dir, cur_dir)
            folders = []
            for file in os.listdir(path):
                if os.path.isdir(os.path.join(path, file)):
                    folders.append(file)
            return JsonResponse({'code': 1, 'folders': folders, 'cur_dir': cur_dir.strip('\\'), 'folder': folder})
        elif state == 'move':
            try:
                move_cur_dir = request.POST.get('move_cur_dir', '')
                move_dir = json.loads(request.POST.get('move_dir', '[]'))
                move_file = json.loads(request.POST.get('move_file', '[]'))
                for move in move_dir:
                    move = os.path.split(move)[1]
                    old_path = os.path.join(nas_dir, cur_dir, move)
                    move_path = os.path.join(nas_dir, move_cur_dir, move)
                    old_sql_path = os.path.join(cur_dir, move)
                    new_sql_path = os.path.join(move_cur_dir, move)

                    if old_path == move_path:
                        return JsonResponse({'code': 0, 'msg': '移动的目标路径和原路径一致，不可进行操作'})
                    if old_path in move_path:
                        return JsonResponse({'code': 0, 'msg': '不可以将父级文件夹移动到子级文件夹中'})
                    shutil.move(old_path, move_path)
                    code, msg = NASSqlite3(request.session.get('user')).NAS_dir_rename_sql(old_sql_path, new_sql_path)
                    if not code:
                        return JsonResponse({'code': 0, 'msg': msg})

                for move in move_file:
                    base = NASSqlite3(request.session.get('user')).NAS_select_savename(move[0], cur_dir)
                    file_path = os.path.join(nas_dir, cur_dir, base)
                    move_path = os.path.join(nas_dir, move_cur_dir, base)
                    code, msg = NASSqlite3(request.session.get('user')).NAS_file_repath_sql(move[0], cur_dir,
                                                                                            move_cur_dir)
                    shutil.move(file_path, move_path)
                    if not code:
                        return JsonResponse({'code': 0, 'msg': msg})
                return JsonResponse({'code': 1, 'msg': '移动完成'})

            except Exception as e:
                return JsonResponse({'code': 0, 'msg': str(e)})
        elif state == 'copy':
            move_cur_dir = request.POST.get('move_cur_dir', '')
            move_dir = json.loads(request.POST.get('move_dir', '[]'))
            move_file = json.loads(request.POST.get('move_file', '[]'))
            for move in move_dir:
                move = os.path.split(move)[1]
                old_path = os.path.join(nas_dir, cur_dir, move)
                move_path = os.path.join(nas_dir, move_cur_dir, move)
                if old_path == move_path:
                    return JsonResponse({'code': 0, 'msg': '复制的目标路径和原路径一致，不可进行操作'})
                # print(old_path.strip(nas_dir), move_path.strip(nas_dir))
                code, msg = NASSqlite3(request.session.get('user')).NAS_copy_insert_sql(
                    old_path.strip(nas_dir), move_path.strip(nas_dir), True)
                if code:
                    for dirpath, dirnames, filenames in os.walk(old_path):
                        for file in filenames:
                            file_path = os.path.join(dirpath, file)
                            new_path = file_path.replace(old_path, move_path)
                            if not os.path.isdir(os.path.split(new_path)[0]):
                                os.makedirs(os.path.split(new_path)[0])
                            shutil.copy(file_path, new_path)

            for move in move_file:
                base = NASSqlite3(request.session.get('user')).NAS_select_savename(move[0], cur_dir)
                old_path = os.path.join(nas_dir, cur_dir, base)
                move_path = os.path.join(nas_dir, move_cur_dir, base)
                code, msg = NASSqlite3(request.session.get('user')).NAS_copy_insert_sql(cur_dir,
                                                                                        move_path.strip(nas_dir), False,
                                                                                        cur_dir)
                if not code:
                    return JsonResponse({'code': 0, 'msg': msg})
                shutil.copy(old_path, move_path)
            return JsonResponse({'code': 1, 'msg': '复制完成'})
        return JsonResponse({'code': 0, 'msg': '请求错误，请刷新后重试'})


# def _NAS_Upload(request):
#     if request.method == 'GET':
#         return JsonResponse({'code': 0, 'msg': '请求错误'})
#     else:
#         file = request.FILES.get('file', None)
#         if file and file.name == request.POST.get('filename') and str(file.size) == request.POST.get('filesize'):
#             chunks = file.chunks()
#             filepath = request.POST.get('filepath', '')
#             filetype = return_file_type(file.name)
#             fileclass = 'other'
#             file_state = 'normal'
#             filepath = filepath.rsplit('/', 1)[0]
#             filepath = filepath.replace('/', '\\')
#             cur_dir = request.POST.get('cur_dir', None)
#             if not os.path.exists(os.path.join(request.session.get('nas_dir'), cur_dir)):
#                 return redirect('/NAS/')
#             save_path = os.path.join(request.session.get('nas_dir'), cur_dir, filepath).replace('&', '-')
#             if not os.path.isdir(save_path):
#                 os.makedirs(save_path)
#             save_path, file_name, save_name = name_nodup(save_path, file.name)
#             save_file = open(save_path, 'wb')
#             for chunk in chunks:
#                 save_file.write(chunk)
#             save_file.close()
#
#             # if num:
#             #     file.name = file.name[::-1].replace('.', '.){}('.format(num), 1)[::-1]
#             #     save_name += '({})'.format(num)
#
#             code, result = NASSqlite3(request.session.get('user')).NAS_file_insert_sql(file_name, save_name,
#                                                                                        os.path.join(cur_dir,
#                                                                                                     filepath).strip(
#                                                                                            '\\'),
#                                                                                        file.size, filetype, fileclass,
#                                                                                        get_time(), file_state)
#             return JsonResponse({'code': 1, 'msg': '上传成功'})
#         return JsonResponse({'code': 0, 'msg': '上传文件不一致，请刷新网页后重试'})

# 用户账户操作
def NAS_user_action(request):
    if request.method == 'GET':
        state = request.GET.get('state')
        # 退出登录
        if state == 'close':
            request.session.delete()
            return JsonResponse({'code': 1, 'url': '/login/'})
    else:
        pass


def NAS_Upload(request):
    if request.method == "GET":
        # 获取前端传来的数据
        HASH = request.session.get('user') + '_' + request.GET.get('HASH', 'None')
        state = request.GET.get('state', None)
        filename = request.GET.get('filename', '').replace('&', '!')
        filepath = request.GET.get('filepath', '')
        filepath = filepath.rsplit('/', 1)[0].replace('/', '\\')
        cur_dir = request.GET.get('cur_dir', '')
        # 获取存储到数据库中的文件的类型
        filetype = return_file_type(filename)
        nas_path = os.path.join(cur_dir, filepath).strip('\\')
        fileclass = 'other'
        file_state = 'normal'
        # 判断是否有该目录，如果没有则跳转至首目录
        if not os.path.exists(os.path.join(request.session.get('nas_dir'), cur_dir)):
            return redirect('/NAS/')
        if HASH and state:
            # 拼接缓存文件路径
            HASH_path = os.path.join(settings.BUFFER_DIR, HASH)

            # 获取前端传过来的文件名的base64值
            path = os.path.join(request.session.get('nas_dir'), cur_dir, filepath)
            # 对文件名和base64名称进行重复校验，获取校验后的名称
            save_path, file_name, save_name = name_nodup(path, filename, settings.IS_NODUP, settings.IS_BASE64_ENCRYPT)
            # 返回已存在缓存文件的文件大小
            if state == 'get_exist_size':
                upload_size = request.GET.get('filesize')
                result = NASSqlite3(request.session.get('user')).NAS_select_user_capacity()
                if int(upload_size) + result[0] > result[1]:
                    return JsonResponse({'code': 0, 'msg': '可用空间不足，无法上传'})
                exist_size = 0
                if os.path.exists(HASH_path):
                    exist_size = os.path.getsize(HASH_path)
                return JsonResponse(
                    {'code': 1, 'msg': '请求成功', 'size': exist_size, 'file_name': file_name, 'save_name': save_name})
            # 将文件拷贝至指定文件夹，并校验后存入数据库中
            elif state == 'end':
                # 判断是否有该目录，如果没有，则批量创建

                if not os.path.isdir(path):
                    os.makedirs(path)
                # 将文件复制到指定目录下
                isex = os.path.exists(save_path)

                shutil.move(HASH_path, save_path)
                file_size = os.path.getsize(save_path)

                if not settings.IS_NODUP and isex:
                    code, result = NASSqlite3(request.session.get('user')).NAS_file_update_sql(file_name, save_name,
                                                                                               nas_path,
                                                                                               file_size, filetype,
                                                                                               fileclass,
                                                                                               get_time(),
                                                                                               file_state)
                else:
                    code, result = NASSqlite3(request.session.get('user')).NAS_file_insert_sql(file_name, save_name,
                                                                                               nas_path,
                                                                                               file_size, filetype,
                                                                                               fileclass,
                                                                                               get_time(),
                                                                                               file_state)

                result = NASSqlite3(request.session.get('user')).NAS_update_user_capacity(file_size)
                print('result：', result)
                if code == 1:
                    return JsonResponse({'code': 1, 'msg': '文件上传成功', 'save_path': save_path})
                return JsonResponse({'code': 0, 'msg': '无法插入数据库，错误代码：{}！'.format(result)})
        if state == 'touch_null_file':
            # 拼接文件路径
            path = os.path.join(request.session.get('nas_dir'), cur_dir, filepath)
            if not os.path.isdir(path):
                os.makedirs(path)
            # 对文件名和base64名称进行重复校验，获取校验后的名称
            save_path, file_name, save_name = name_nodup(path, filename, settings.IS_NODUP, settings.IS_BASE64_ENCRYPT)
            with open(save_path, 'a+') as f:
                f.write('')
            code, result = NASSqlite3(request.session.get('user')).NAS_file_insert_sql(file_name, save_name,
                                                                                       nas_path,
                                                                                       0, filetype,
                                                                                       fileclass,
                                                                                       get_time(),
                                                                                       file_state)
            if code == 1:
                return JsonResponse({'code': 1, 'msg': '文件上传成功', 'save_path': save_path})
            return JsonResponse({'code': 0, 'msg': '无法插入数据库，错误代码：{}！'.format(result)})
        return JsonResponse({'code': 0, 'msg': '请求错误，请刷新网页后重试'})
    else:
        # 获取前端传过来的数据
        HASH = request.session.get('user') + '_' + request.POST.get('HASH', 'None')
        state = request.POST.get('state', None)
        data = request.FILES.get('file', None)
        # 判断磁盘可用容量是否大于上传数据的容量
        free_space = getLocalSpace(settings.NAS_DIR)
        if free_space < data.size * 2:
            return JsonResponse({'code': 0, 'msg': '当前NAS服务器磁盘可用容量不足，请联系管理员'})
        if HASH and state and data:
            HASH_path = os.path.join(settings.BUFFER_DIR, HASH)
            # 判断是否存在缓存目录
            if not os.path.exists(settings.BUFFER_DIR):
                os.mkdir(settings.BUFFER_DIR)
            # 将文件写入缓存目录下
            if state == 'upload':
                file = open(HASH_path, 'ab')
                for chunk in data.chunks():
                    file.write(chunk)
                file.close()
                return JsonResponse({'code': 1, 'msg': '上传成功', 'size': data.size})
        return JsonResponse({'code': 0, 'msg': '请求错误，请刷新网页后重试'})


# 下载文件方法
def NAS_download_file(request):
    if request.method == 'GET':
        return redirect('/NAS/')
    else:
        dirlist = json.loads(request.POST.get('download_dir', '[]'))
        filelist = json.loads(request.POST.get('download_file', '[]'))
        state = request.POST.get('state', None)
        zip_path = request.POST.get('zip_path', None)
        if zip_path:
            print(zip_path.strip(request.session.get('nas_dir')))
        file_name = request.POST.get('file_name', None)
        if state == 'get_size':
            size = 0
            for dir in dirlist:
                dir = os.path.join(request.session.get('nas_dir'), dir)
                for path, dirname, filename in os.walk(dir):
                    for file in filename:
                        p = os.path.join(path, file)
                        size += os.path.getsize(p)
            for file_name, date, file_size, path in filelist:
                save_name = NASSqlite3(request.session.get('user')).NAS_select_savename(file_name, path)
                file_path = os.path.join(request.session.get('nas_dir'), path, save_name)
                size += os.path.getsize(file_path)
            return JsonResponse({'code': 1, 'size': size})
        elif state == 'touch_zip':
            zip_list = []
            for path in dirlist:
                zip_list.append(os.path.join(request.session.get('nas_dir'), path.replace('/', '\\')))
            for file_name, date, size, path in filelist:
                save_name = NASSqlite3(request.session.get('user')).NAS_select_savename(file_name, path)
                zip_list.append(os.path.join(request.session.get('nas_dir'), path.replace('/', '\\'), save_name))
            zip_path = os.path.join(settings.BUFFER_DIR, get_random() + '.zip')
            compress(zip_list, zip_path, request.session.get('user'), request.session.get('nas_dir'))
            return JsonResponse({'code': 1, 'msg': '文件压缩完成！', 'zip_path': zip_path, 'file_name': 'NAS_download.zip'})
        elif state == 'download_one':
            file_name = filelist[0][0]
            save_path = filelist[0][3]
            data = NASSqlite3(request.session.get('user')).NAS_file_select_sql(file_name=file_name, file_path=save_path)
            base_name = data[1][0][1]
            file_path = os.path.join(request.session.get('nas_dir'), save_path, base_name)
            if os.path.isfile(file_path):
                return JsonResponse({'code': 1, 'msg': '文件存在！', 'zip_path': file_path, 'file_name': file_name})
            else:
                return JsonResponse({'code': 0, 'msg': '文件不存在！'})
        elif state == 'download':
            if os.path.isfile(zip_path):
                file = open(zip_path, 'rb')
                response = FileResponse(file)
                response['Content-Type'] = 'application/octet-stream'
                response['Content-Disposition'] = 'attachment;filename="{}"'.format(
                    escape_uri_path(file_name.encode('utf-8')))
                return response
            else:
                return JsonResponse({'code': 0, 'msg': '文件不存在！'})

        elif state == 'remove_zip':
            if os.path.isfile(zip_path):
                while 1:
                    time.sleep(5)
                    try:
                        os.remove(zip_path)
                        # 循环判断是否有超过一天没有删除的文件
                        for file in os.listdir(settings.BUFFER_DIR):
                            file_path = os.path.join(settings.BUFFER_DIR, file)
                            ctime = os.path.getctime(file_path)
                            if time.time() - ctime > 60 * 60 * 24:
                                if '.zip' in file:
                                    os.remove(file_path)
                        return JsonResponse({'code': 1})
                    except Exception as e:
                        pass
            return JsonResponse({'code': 1})
        else:
            return redirect('/NAS/')


# 删除文件方法
def NAS_delete(request):
    if request.method == 'POST':
        del_dir = json.loads(request.POST.get('del_dir', '["",""]'))
        del_file = json.loads(request.POST.get('del_file', '["",""]'))
        code, res = del_file_fun(del_dir, del_file, request.session.get('user'), request.session.get('nas_dir'))
        return JsonResponse({'code': code, 'msg': res})


def del_file_fun(del_dir, del_file, user, nas_dir):
    try:
        del_dir_count = 0
        del_file_count = 0
        # 循环删除目录以及目录下面的文件，并删除数据库记录
        for dir_ in del_dir:
            dir_ = os.path.join(nas_dir, dir_)
            dir_list = dir_[len(nas_dir):].replace(r'/', '\\').strip('\\')
            code, res = NASSqlite3(user).NAS_delete_file('', '', '', dir_list)
            if code == 0:
                raise res

            def fun(path, del_dir_count, del_file_count):
                for f in os.listdir(path):
                    d = os.path.join(path, f)
                    if os.path.isdir(d):
                        dir_list = d[len(nas_dir):].replace(r'/', '\\').strip('\\')
                        code, res = NASSqlite3(user).NAS_delete_file('', '', '', dir_list)
                        del_dir_count, del_file_count = fun(d, del_dir_count, del_file_count)
                        os.rmdir(d)
                        del_dir_count += 1
                    else:
                        capacityUsed = os.path.getsize(d)
                        capacityUsed = ~capacityUsed + 1
                        NASSqlite3(user).NAS_update_user_capacity(capacityUsed)
                        os.remove(d)
                        del_file_count += 1
                return del_dir_count, del_file_count

            del_dir_count, del_file_count = fun(dir_, del_dir_count, del_file_count)
            os.rmdir(dir_)
            del_dir_count += 1
        # 循环删除当前目录下的文件，并删除数据库记录
        for file in del_file:
            file_name = file[0]
            file_upload_date = file[1]
            file_size = file[2]
            sql_path = file[3]

            code, res = NASSqlite3(user).NAS_delete_file(file_name, file_upload_date, file_size,
                                                         sql_path)
            capacityUsed = os.path.getsize(os.path.join(nas_dir, sql_path, res))
            capacityUsed = ~capacityUsed + 1
            NASSqlite3(user).NAS_update_user_capacity(capacityUsed)
            if code == 1:
                file_path = os.path.join(nas_dir, sql_path, res)
                if os.path.isfile(file_path):
                    os.remove(file_path)
                del_file_count += 1
            else:
                raise res
        return [1, '删除成功，共删除了{}个文件夹，{}个文件'.format(del_dir_count, del_file_count)]
    except Exception as e:
        return [0, str(e)]


# 返回服务器端请求的图片
def get_image(request):
    if request.method == 'GET':
        # 获取前端发送的数据
        filename = request.GET.get('filename', None)
        cur_dir = request.GET.get('cur_dir', '')
        compress = request.GET.get('compress', True)
        if filename:
            save_path = os.path.join(request.session.get('nas_dir'), cur_dir, filename)
            # print(save_path)
            if not os.path.exists(save_path):
                return JsonResponse({'code': 0, 'msg': '请求的文件不存在'})
            image_data = open(save_path, "rb").read()
            if compress is True:
                # 预览图压缩处理， 压缩为200px宽，高度自适应
                from PIL import Image
                from io import BytesIO
                by = BytesIO()
                im = Image.open(save_path)
                x, y = im.size
                ratio = x / 200
                # 修改图片大小
                image_data = im.resize((int(x / ratio), int(y / ratio)), Image.ANTIALIAS)
                # 将图片转化为Bytes
                image_data.save(by, format('PNG'))
                image_data = by.getvalue()
            return HttpResponse(image_data, content_type="image/png")
        else:
            return JsonResponse({'code': 0, 'msg': '请求的文件不存在'})


# 读取视频文件，使用yield分块输出
def file_iterator(file_name, chunk_size=8192, offset=0, length=None):
    with open(file_name, "rb") as f:
        f.seek(offset, os.SEEK_SET)
        remaining = length
        while True:
            bytes_length = chunk_size if remaining is None else min(remaining, chunk_size)
            data = f.read(bytes_length)
            if not data:
                break
            if remaining:
                remaining -= len(data)
            yield data


# 将视频文件以流媒体的方式响应
def stream_video(request, path):
    range_header = request.META.get('HTTP_RANGE', '').strip()
    range_re = re.compile(r'bytes\s*=\s*(\d+)\s*-\s*(\d*)', re.I)
    range_match = range_re.match(range_header)
    size = os.path.getsize(path)
    # content_type, encoding = mimetypes.guess_type(path)
    # content_type = content_type or 'application/octet-stream'
    content_type = 'application/octet-stream'
    if range_match:
        first_byte, last_byte = range_match.groups()
        first_byte = int(first_byte) if first_byte else 0
        last_byte = first_byte + 1024 * 1024 * 8  # 8M 每片,响应体最大体积
        if last_byte >= size:
            last_byte = size - 1
            length = last_byte - first_byte + 1
            resp = StreamingHttpResponse(file_iterator(path, offset=first_byte, length=length), status=206,
                                         content_type=content_type)
            resp['Content-Length'] = str(length)
            resp['Content-Range'] = 'bytes %s-%s/%s' % (first_byte, last_byte, size)
        else:
            # 不是以视频流方式的获取时，以生成器方式返回整个文件，节省内存
            resp = StreamingHttpResponse(FileWrapper(open(path, 'rb')), content_type=content_type)
            resp['Content-Length'] = str(size)
        resp['Accept-Ranges'] = 'bytes'
        return resp


# 将读取出来的excel内容格式化为html
def excel_to_html(excel_list):
    html = ''
    for sheet in excel_list:
        sheet_name = list(sheet.keys())[0]
        sheet_text = '<h1>{0}</h1>\n<table class="table table-bordered table-hover {0}">\n\t<tbody>'.format(sheet_name)
        value_list = sheet[sheet_name]
        for row, col_list in enumerate(value_list):
            tr = '<tr>\n'
            for col, value in enumerate(col_list):
                td = '<td>{}</td>'.format(value)
                tr += td
            tr += '</tr>'
            sheet_text += tr
        sheet_text += '\t</tbody>\n</table>\n'
        html += sheet_text
    return html


# 读取excel内容
def read_excel(filepath, filename):
    pythoncom.CoInitialize()
    save_path = os.path.join(settings.BUFFER_DIR, 'DOC', filename)
    shutil.copy(filepath, save_path)
    filepath = save_path
    if filename[-4:] == '.xls':
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(filepath)
        wb.SaveAs(filepath + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        filepath += "x"
    excel = openpyxl.open(filepath, data_only=True)
    excel_list = []
    merged_cell_dict = {}
    for sheet in excel.sheetnames:
        wb = excel[sheet]
        merged_list = []
        max_rows = wb.max_row
        max_cols = wb.max_column
        for merged in wb.merged_cell_ranges:
            left, right = str(merged).split(':')
            re_left = re.match('[A-Z]+', left).group()
            re_right = re.match('[A-Z]+', right).group()
            merged = '{}:{}'.format(
                left.replace(re_left, str(column_index_from_string(re_left)) + '-'),
                right.replace(re_right, str(column_index_from_string(re_right)) + '-'))
            merged_list.append(merged)
        merged_cell_dict[sheet] = merged_list
        sheet_dict = {}
        sheet_list = []
        for row in range(1, max_rows+1):
            row_list = []
            for col in range(1, max_cols+1):
                val = wb.cell(row, col).value
                if val:
                    row_list.append(val)
                else:
                    row_list.append('')
            sheet_list.append(row_list)
            sheet_dict[sheet] = sheet_list
        excel_list.append(sheet_dict)
    return excel_list, merged_cell_dict


# 将输入的合并单元格格式化为列表形式，并且行和列进行分开
def sort_merged_dict(merged):
    row_merged_dict = {}
    col_merged_dict = {}
    for key in merged.keys():
        value = merged[key]
        row_merged_dict[key] = []
        col_merged_dict[key] = []
        for val in value:
            left_col = val.split(':')[0].split('-')[0]
            left_row = val.split(':')[0].split('-')[1]
            rig_col = val.split(':')[1].split('-')[0]
            rig_row = val.split(':')[1].split('-')[1]
            if left_row == rig_row:
                row_merged_dict[key].append({left_row: [left_col, rig_col]})
                # print('行合并：将第{}行从第{}列到第{}列进行合并'.format(left_row, left_col, rig_col))
            if left_col == rig_col:
                col_merged_dict[key].append({left_col: [left_row, rig_row]})
                # print('列合并：将第{}列从第{}行到第{}行进行合并'.format(left_col, left_row, rig_row))
    return row_merged_dict, col_merged_dict


# 返回服务器端请求的预览文件信息
def online_preview(request):
    user = request.session.get('user')
    nas_dir = request.session.get('nas_dir')
    if request.method == 'GET':
        cur_dir = request.GET.get('cur_dir', '')
        filename = request.GET.get('filename', None)
        file_type = request.GET.get('type', None)
        if filename and file_type:
            result = NASSqlite3(user).NAS_select_savename(filename, cur_dir)
            if not os.path.exists(os.path.join(settings.BUFFER_DIR, 'DOC')):
                os.mkdir(os.path.join(settings.BUFFER_DIR, 'DOC'))
            if result:
                path = os.path.join(nas_dir, cur_dir, result)
            else:
                return JsonResponse({'code': 2, 'msg': '请求错误'})
            if not request.GET.get('html', None):
                for file in os.listdir(os.path.join(settings.BUFFER_DIR, 'DOC')):
                    if file != filename:
                        os.remove(os.path.join(os.path.join(settings.BUFFER_DIR, 'DOC', file)))
                if file_type == 'pdf':
                    shutil.copy(path, os.path.join(settings.BUFFER_DIR, 'DOC', filename))
                    url = '/static/buffer/DOC/{}'.format(filename)
                elif file_type == 'ppt':
                    save_name = filename.replace('.pptx', '.pdf').replace('.ppt', '.pdf')
                    save_path = os.path.join(os.path.join(settings.BUFFER_DIR, 'DOC', save_name))
                    res = ppt_to_pdf(path, save_path)
                    if res[0] == 1:
                        url = '/static/buffer/DOC/{}'.format(save_name)
                        file_type = 'pdf'
                    else:
                        return JsonResponse({'code': res[0], 'msg': res[1]})
                elif file_type == 'word':
                    from pydocx import PyDocX
                    shutil.copy(path, os.path.join(settings.BUFFER_DIR, 'DOC', filename))
                    docx = os.path.join(settings.BUFFER_DIR, 'DOC', filename)
                    word = PyDocX.to_html(docx)
                    return HttpResponse(word)
                elif file_type == 'excel':
                    excel_list, merged_cell_dict = read_excel(path, filename)
                    row_merged_dict, col_merged_dict = sort_merged_dict(merged_cell_dict)
                    html = excel_to_html(excel_list)
                    return render(request, '{}.html'.format(file_type),
                                  {'html': html,
                                   'row_merged_dict': json.dumps(row_merged_dict),
                                   'col_merged_dict': json.dumps(col_merged_dict)})
                elif file_type == 'txt' or file_type == 'conf' or file_type == 'bat':
                    file_type = 'txt'
                    save_path = os.path.join(settings.BASE_DIR, 'templates', 'txt.html')
                    url = ''
                    res = txt_to_html(path, save_path)
                    if res[0] == 1:
                        return render(request, '{}.html'.format(file_type))
                else:
                    url = '/online_preview/?cur_dir={}&filename={}&type={}'.format(cur_dir, filename, file_type)

                return render(request, '{}.html'.format(file_type),
                              {'url': url})

            file = stream_video(request, path)
            return file
        return JsonResponse({'code': 0, 'msg': '请求错误'})
    else:

        return JsonResponse({'code': 0, 'msg': '请求错误'})


# ppt和pptx文件转换
def ppt_to_pdf(filepath, savepath):
    pythoncom.CoInitialize()
    if os.path.isfile(savepath):
        return 2, '已经转换了'
    p = gencache.EnsureDispatch("PowerPoint.Application")
    try:
        ppt = p.Presentations.Open(filepath, False, False, False)
    except Exception as e:
        return os.path.split(filepath)[1], "转化失败,失败原因%s" % e
    ppt.ExportAsFixedFormat(savepath, 2, PrintRange=None)
    p.Quit()
    return 1, '转换成功'


# 文本文件转为html
def txt_to_html(filepath, savepath):
    try:
        with open(filepath, 'rb') as f:
            data = f.read()
            code = chardet.detect(data)['encoding']
        if code != 'utf-8':
            with open(filepath, "r", encoding=code) as file:
                file = file.read()
            filepath = os.path.join(settings.BUFFER_DIR, 'DOC', 'to_txt.txt')
            with open(filepath, "w", encoding='utf-8') as f:
                f.write(file)
        with open(filepath, "r", encoding='utf-8') as file:
            contents = file.readlines()
        with open(savepath, "w", encoding='utf-8') as e:
            txt_html = ''
            e.write("<style>*{font-size: 18px}</style><title>文本文件预览</title>")
            for lines in contents:
                e.write("<pre>" + lines + "</pre>\n")
                txt_html += "<pre>" + lines + "</pre>\n"
        return 1, " style='font-size: 18px'"
    except Exception as e:
        return 0, '转化失败，错误代码{}'.format(str(e))


# 返回服务器请求的pdf文件信息
def get_pdf(request):
    user = request.session.get('user')
    nas_dir = request.session.get('nas_dir')
    if request.method == 'GET':
        cur_dir = request.GET.get('cur_dir', '')
        filename = request.GET.get('filename', None)
        if not request.GET.get('html', None):
            return render(request, 'video.html',
                          {'url': '/online_preview/?cur_dir={}&filename={}'.format(cur_dir, filename)})
        if filename:
            result = NASSqlite3(user).NAS_select_savename(filename, cur_dir)
            if result:
                path = os.path.join(nas_dir, cur_dir, result)
                file = stream_video(request, path)
                return file
        return JsonResponse({'code': 0, 'msg': '请求错误'}, charset='utf-8')
    else:
        pass
    return JsonResponse({'code': 0, 'msg': '访问错误'})


# 返回文件类型
def return_file_type(file_name):
    filetype = 'unknown'
    if file_name.find('.'):
        suffix = file_name.split('.')[-1].lower()
        for key in settings.NAS_FILE_TYPE.keys():
            val = settings.NAS_FILE_TYPE[key]
            if suffix in val:
                filetype = key
    return filetype


# 返回当前目录下的文件夹
def get_folders(nas_dir, cur_dir):
    root_dir = os.listdir(os.path.join(nas_dir, cur_dir))
    folders = []
    for fol in root_dir:
        fol_path = os.path.join(nas_dir, cur_dir, fol)
        if os.path.isdir(fol_path):
            fol_info = {'name': fol,
                        'fol_path': fol_path.replace(nas_dir + '\\', ''),
                        'ctime': get_time(os.path.getctime(fol_path), cn=False),
                        'mtime': get_time(os.path.getmtime(fol_path), cn=False)}
            folders.append(fol_info)
    return folders


# 格式化文件大小
def file_size_format(file_size):
    if type(file_size) != int:
        return file_size
    if file_size > 1024 * 1024 * 1024 * 1024:
        file_size = '{}TB'.format(round(file_size / (1024 * 1024 * 1024 * 1024), 1))
    elif file_size > 1024 * 1024 * 1024:
        file_size = '{}GB'.format(round(file_size / (1024 * 1024 * 1024), 1))
    elif file_size > 1024 * 1024:
        file_size = '{}MB'.format(round(file_size / (1024 * 1024), 1))
    elif file_size > 1024:
        file_size = '{}KB'.format(round(file_size / 1024, 1))

    return file_size


# 将文件添加到压缩文件中
def compress(sou_files_path, des_files_path, user, nas_dir):
    """
    压缩文件生成.zip 格式文件
    :param sou_files_path: 须要压缩的文件夹
    :param des_files_path: 存放压缩文件的地址
    :param user 接收用户名，用于区分属于谁的数据
    :return:
    """
    f = zipfile.ZipFile(des_files_path, 'w', zipfile.ZIP_DEFLATED)

    for z in sou_files_path:
        if os.path.isdir(z):
            root = z.split('\\')[-1]
            root_dir = z.replace(root, '')
            for dirpath, dirnames, filenames in os.walk(z):
                for filename in filenames:
                    name = NASSqlite3(user).NAS_select_filename(filename, dirpath.replace(nas_dir, '').strip('\\'))
                    save_file = os.path.join(dirpath.replace(root_dir, ''), name)
                    f.write(os.path.join(dirpath, filename), save_file)
        else:
            name = NASSqlite3(user).NAS_select_filename(os.path.split(z)[1], os.path.split(z)[0].replace(nas_dir, '')
                                                        .strip('\\'))
            f.write(z, name)
    f.close()
    return f


# 获取随机的字符串，由数字和大小写字母组成
def get_random(length=10):
    import random
    s = ''
    for i in range(length):
        a = [random.randint(48, 57), random.randint(65, 90), random.randint(97, 122)]
        s += str(chr(a[random.randint(0, 2)]))
    return s


# 判断文件有没有重复，如果重复则返回一个不重复的文件名
def name_nodup(file_path, file_name='', nodup=True, base_encrypt=True):
    num = 0
    while 1:
        if base_encrypt:
            save_name = base64.b64encode(file_name.encode('utf-8')).decode('utf-8').replace('+', ',').replace('/', '.')
        else:
            save_name = file_name
        if nodup:
            if os.path.isfile(os.path.join(file_path, save_name)):

                if file_name.count('({})'.format(num)) > 0:
                    file_name = file_name.replace('({})'.format(num), '({})'.format(num + 1), 1)
                else:
                    if file_name.count('.') > 0:
                        file_name = file_name[::-1].replace('.', '.){}('.format(num + 1), 1)[::-1]
                    else:
                        file_name += '({})'.format(num + 1)
                num += 1
            else:
                break
        else:
            break
    save_path = os.path.join(file_path, save_name)
    if len(save_path) > 257:
        outrange = 257 - len(save_path)
        save_name = save_name[:outrange]
        save_path = os.path.join(file_path, save_name)
    return save_path, file_name, save_name


# 校验名称是否能够创建文件夹
def validity(text):
    return not re.search(r'[\\/:*?"<>|]', text)


# 格式化输出时间
def get_time(seconds=None, ishour=True, cn=False):
    if not seconds:
        seconds = time.time()
    local = time.localtime(seconds)
    year = local.tm_year
    mon = local.tm_mon
    day = local.tm_mday
    hour = local.tm_hour
    min = local.tm_min
    sec = local.tm_sec
    if ishour:
        if cn:
            return '{}年{:0>2d}月{:0>2d}日 {:0>2d}时{:0>2d}分{:0>2d}秒'.format(year, mon, day, hour, min, sec)
        return '{}-{:0>2d}-{:0>2d} {:0>2d}:{:0>2d}:{:0>2d}'.format(year, mon, day, hour, min, sec)
    if cn:
        return '{}年{:0>2d}月{:0>2d}日'.format(year, mon, day)
    return '{}-{:0>2d}-{:0>2d}'.format(year, mon, day)


# 获取磁盘剩余空间
def getLocalSpace(folder):
    """
    获取磁盘剩余空间
    :param folder: 磁盘路径 例如 D:\\
    :return: 剩余空间 单位 G
    """
    folderTemp = folder
    if not os.path.exists(folderTemp):
        folderTemp = os.getcwd()
    if platform.system() == 'Windows':
        free_bytes = ctypes.c_ulonglong(0)
        ctypes.windll.kernel32.GetDiskFreeSpaceExW(ctypes.c_wchar_p(folderTemp), None, None, ctypes.pointer(free_bytes))
        return free_bytes.value
    else:
        st = os.statvfs(folderTemp)
        return st.f_bavail * st.f_frsize / 1024


def test(request):
    return render(request, 'test.html')




